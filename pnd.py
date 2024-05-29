import sqlite3
from openpyxl import load_workbook
from datetime import datetime
from openpyxl.styles import NamedStyle


# Подключение к базе данных
conn = sqlite3.connect('news_data.db')
conn.row_factory = sqlite3.Row  # Позволяет обращаться к столбцам по имени

# Выбор строк, где company не равно '-' и added_to_xls равно FALSE
query = "SELECT rowid, link, title, text, date_published1, company FROM news WHERE company != '-' AND added_to_xls IS TRUE"
cursor = conn.execute(query)
new_data = cursor.fetchall()  # Получаем данные в виде списка словарей

# Загрузка xlsx-файла
wb = load_workbook('news.xlsx')
sheet = wb['Свод новостей']  # Получаем нужный лист

# Определение начальной строки для записи
current_row = sheet.max_row + 1  # Находим первую пустую строку

# Запись данных и сбор rowid для обновления
rowids_to_update = []

style_name = 'datetime'
if style_name not in wb.named_styles:
    date_style = NamedStyle(name=style_name, number_format='MM/DD/YYYY')
    wb.add_named_style(date_style)

for row in new_data:
    d = row['date_published1']
    sheet[f'B{current_row}'] = row['title']
    sheet[f'C{current_row}'] = row['company']
    sheet[f'D{current_row}'] = d[5:7]
    sheet[f'E{current_row}'] = d[:4]

    date_str = f'{d[5:7]}/{d[8:10]}/{d[:4]}'
    date_obj = datetime.strptime(date_str, '%m/%d/%Y')
    sheet[f'F{current_row}'] = date_obj
    sheet[f'F{current_row}'].style = style_name

    sheet[f'AB{current_row}'] = row['text']
    sheet[f'AC{current_row}'] = row['link']
    sheet[f'AD{current_row}'] = 'Металлоснабжение и сбыт'
    rowids_to_update.append(row['rowid'])  # Сохраняем rowid для обновления
    current_row += 1

# Сохранение изменений в Excel
wb.save('news.xlsx')

# Обновление столбца added_to_xls на TRUE для записанных строк
if rowids_to_update:
    placeholders = ', '.join(['?'] * len(rowids_to_update))
    update_query = f"UPDATE news SET added_to_xls = TRUE WHERE rowid IN ({placeholders})"
    conn.execute(update_query, rowids_to_update)
    conn.commit()

# Закрытие соединения с базой данных
conn.close()
